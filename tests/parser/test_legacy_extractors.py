"""Tests for legacy text extraction helpers."""

from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook

from lightrag.parser.legacy.extractors import extract_text, _extract_xls


_NS_URI = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS = {"main": _NS_URI}


def _inject_cached_value(
    data: bytes, cell_ref: str, cached_value: str | int | float
) -> bytes:
    """Patch a worksheet's XML so ``cell_ref`` carries a cached ``<v>`` value.

    ``openpyxl`` writes formula expressions but never a cached calculated value,
    so we inject one to exercise the ``data_only=True`` read path. String results
    are tagged ``t="str"`` to match how Excel records a text-valued formula.
    """

    root = ET.fromstring(data)
    for cell in root.findall(".//main:c", _NS):
        if cell.attrib.get("r") != cell_ref:
            continue
        if isinstance(cached_value, str):
            cell.set("t", "str")
        value_node = cell.find("main:v", _NS)
        if value_node is None:
            value_node = ET.SubElement(cell, f"{{{_NS_URI}}}v")
        value_node.text = str(cached_value)
        break
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


def _patch_xlsx(
    file_bytes: bytes, injections: dict[str, tuple[str, str | int | float]]
) -> bytes:
    """Rewrite worksheet parts in ``file_bytes`` with cached formula values.

    ``injections`` maps a worksheet part name (e.g. ``xl/worksheets/sheet1.xml``)
    to a ``(cell_ref, cached_value)`` pair.
    """

    source = BytesIO(file_bytes)
    patched = BytesIO()
    with ZipFile(source, "r") as zin, ZipFile(patched, "w", ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename in injections:
                cell_ref, cached_value = injections[item.filename]
                data = _inject_cached_value(data, cell_ref, cached_value)
            zout.writestr(item, data)
    return patched.getvalue()


def _make_xlsx_bytes(*, cached_formula_value: str | int | float | None) -> bytes:
    """Build a minimal single-sheet workbook with one formula cell.

    ``openpyxl`` writes the formula expression, but not a cached calculated
    value. When ``cached_formula_value`` is given we patch the worksheet XML so
    the extractor exercises the ``data_only=True`` path; when it is ``None`` the
    workbook carries no cache and the formula-text fallback path is exercised.
    """

    bio = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["B1"] = "=SUM(A1:A2)"
    wb.save(bio)

    if cached_formula_value is None:
        return bio.getvalue()

    return _patch_xlsx(
        bio.getvalue(), {"xl/worksheets/sheet1.xml": ("B1", cached_formula_value)}
    )


def _make_multi_sheet_xlsx_bytes() -> bytes:
    """Two sheets of differing shape, each with a cached formula result.

    Exercises the per-sheet title matching (``wb_formulas[sheet.title]``), the
    cross-view dimension union, and a string-valued cached result.
    """

    bio = BytesIO()
    wb = Workbook()
    numbers = wb.active
    numbers.title = "Numbers"
    numbers["A1"] = 1
    numbers["A2"] = 2
    numbers["B1"] = "=SUM(A1:A2)"

    words = wb.create_sheet("Words")
    words["A1"] = "foo"
    words["B1"] = '=A1&"bar"'
    wb.save(bio)

    return _patch_xlsx(
        bio.getvalue(),
        {
            "xl/worksheets/sheet1.xml": ("B1", 3),
            "xl/worksheets/sheet2.xml": ("B1", "foobar"),
        },
    )


@pytest.mark.offline
def test_extract_text_xlsx_uses_cached_formula_value():
    file_bytes = _make_xlsx_bytes(cached_formula_value=3)

    text = extract_text(file_bytes, "xlsx")

    assert "3" in text
    assert "=SUM(A1:A2)" not in text


@pytest.mark.offline
def test_extract_text_xlsx_falls_back_to_formula_text_when_cache_missing():
    file_bytes = _make_xlsx_bytes(cached_formula_value=None)

    text = extract_text(file_bytes, "xlsx")

    assert "=SUM(A1:A2)" in text


@pytest.mark.offline
def test_extract_text_xlsx_handles_multiple_sheets_and_string_results():
    file_bytes = _make_multi_sheet_xlsx_bytes()

    text = extract_text(file_bytes, "xlsx")

    # Both sheets are emitted, matched to their own formula view by title.
    assert "Sheet: Numbers" in text
    assert "Sheet: Words" in text
    # Numeric cached result preferred over the SUM formula text.
    assert "3" in text
    assert "=SUM(A1:A2)" not in text
    # String cached result preferred over the concatenation formula text.
    assert "foobar" in text
    assert '=A1&"bar"' not in text


# ---------------------------------------------------------------------------
# XLS helpers (xlwt creates .xls BIFF files for test purposes)
# ---------------------------------------------------------------------------


def _make_xls_bytes() -> bytes:
    """Two-sheet workbook with numbers, strings, and a header row."""
    import xlwt  # type: ignore

    wb = xlwt.Workbook(encoding="utf-8")
    ws1 = wb.add_sheet("Numbers")
    ws1.write(0, 0, "A")
    ws1.write(0, 1, "B")
    ws1.write(1, 0, 42)
    ws1.write(1, 1, 3.14)
    ws1.write(2, 0, "hello")
    ws1.write(2, 1, "world")

    ws2 = wb.add_sheet("Data")
    ws2.write(0, 0, "ID")
    ws2.write(0, 1, "Name")
    ws2.write(1, 0, 1)
    ws2.write(1, 1, "Alice")
    ws2.write(2, 0, 2)
    ws2.write(2, 1, "Bob")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# XLS tests
# ---------------------------------------------------------------------------


@pytest.mark.offline
def test_extract_text_xls_basic_extraction():
    """Verify two-sheet .xls extraction produces sheet separators and data."""
    file_bytes = _make_xls_bytes()
    text = _extract_xls(file_bytes)

    assert "Sheet: Numbers" in text
    assert "Sheet: Data" in text
    assert "A\tB" in text
    assert "ID\tName" in text
    # xlwt stores integers as floats in BIFF format
    assert "42.0\t3.14" in text
    assert "hello\tworld" in text
    assert "1.0\tAlice" in text
    assert "2.0\tBob" in text
    assert "====================" in text


@pytest.mark.offline
def test_extract_text_xls_via_dispatch():
    """Verify extract_text dispatches 'xls' suffix to _extract_xls."""
    file_bytes = _make_xls_bytes()
    text = extract_text(file_bytes, "xls")
    assert "Sheet: Numbers" in text
    assert "42.0\t3.14" in text


@pytest.mark.offline
def test_extract_text_xls_empty_cells():
    """Verify empty cells produce empty-string columns (tab separators)."""
    import xlwt

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "left")
    ws.write(0, 2, "right")  # col 1 is empty
    bio = BytesIO()
    wb.save(bio)
    text = _extract_xls(bio.getvalue())

    assert "left" in text
    assert "right" in text
    lines = [l for l in text.split("\n") if "left" in l]
    assert len(lines) == 1
    assert "left\t\tright" in lines[0]


@pytest.mark.offline
def test_extract_text_xls_special_chars():
    """Verify newlines and tabs in cell values are escaped."""
    import xlwt

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "line1\nline2")
    ws.write(0, 1, "col1\tcol2")
    bio = BytesIO()
    wb.save(bio)
    text = _extract_xls(bio.getvalue())

    assert "line1\\nline2" in text
    assert "col1\\tcol2" in text
